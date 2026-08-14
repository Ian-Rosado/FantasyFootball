import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
wb=openpyxl.Workbook()
navy="1F3A5F"; hf=PatternFill("solid",fgColor=navy); hfont=Font(bold=True,color="FFFFFF",size=11)
bd=Border(bottom=Side(style="thin",color="D0D0D0"))
def add(name,path,widths=None,numfmts=None):
    ws=wb.create_sheet(name); rows=list(csv.reader(open(path,encoding="utf-8")))
    for r,row in enumerate(rows,1):
        for c,val in enumerate(row,1):
            cell=ws.cell(r,c)
            if r>1:
                try: val="" if val=="" else (int(val) if val.lstrip("-").isdigit() else float(val))
                except: pass
            cell.value=val
            if r==1: cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal="center")
            else: cell.border=bd
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
    for c in range(1,len(rows[0])+1):
        L=get_column_letter(c)
        w=widths.get(c) if (widths and c in widths) else max(10,min(30,max(len(str(rows[r][c-1])) for r in range(min(len(rows),60)) if c-1<len(rows[r]))+2))
        ws.column_dimensions[L].width=w
    if numfmts:
        for c,fmt in numfmts.items():
            for r in range(2,len(rows)+1): ws.cell(r,c).number_format=fmt

ab=wb.active; ab.title="About"; ab.column_dimensions["A"].width=114
L=[("Gridiron Grind — League History & Analysis",True,16),
 ("Fleaflicker league #78455 (2009–2022) + ESPN league #216270753 (2023–2025)  •  12-team, 2-keeper  •  refreshed Aug 2026",False,10),("",False,10),
 ("Managers tracked by owner account so records follow name/franchise changes. Reg season 13 games (2009–2020), 14 (2021–2025).",False,11),("",False,10),
 ("CHAMPIONS: taken from the actual title game each year. Fleaflicker's 'postseason rank' field is unreliable and is NOT used.",False,11),("",False,10),
 ("TABS",True,12),
 ("Manager_AllTime — records, titles, runner-ups, title-game apps, #1 seeds, playoff apps, best/avg regular-season rank.",False,11),
 ("Champions — title game champion, runner-up, score, #1 seed each season.",False,11),
 ("Standings — every team-season (record, PF/PA, seed, playoff W-L, draft slot).",False,11),
 ("Luck_AllTime / Luck_BySeason — actual vs. expected wins (all-play).   H2H_Matrix — all-time head-to-head.",False,11),
 ("Trades / Trades_ByManager — all completed trades and per-manager counts.",False,11),
 ("Draft_Fleaflicker — snake-draft boards 2013,2014,2016–2022.",False,11),
 ("Draft_ESPN_Auction — auction boards 2023–2025 with $ bid amounts and keeper flags (keepers only flagged by ESPN in 2025).",False,11),
 ("FAAB_Claims — every FAAB waiver win 2021–2025 with winning bid, # of bidders, and top losing bid.",False,11),
 ("FAAB_ByManager — FAAB spend/claims/avg/max per manager per season and all-time.",False,11),("",False,10),
 ("NOTES",True,12),
 ("• ESPN owners mapped to Fleaflicker managers (maxdawkins = madmax101010, Michael Oxmall = Brotesk).",False,11),
 ("• FAAB began in 2021; all five seasons (2021–2025) are complete.   • Full weekly game log (1,483 games) available on request.",False,11),
 ("• 2015 draft and any pre-2013 drafts were not recorded on Fleaflicker.",False,11),("",False,10),
 ("DRAFT PREP 2026 (one-keeper year)",True,12),
 ("Current_Rosters - end-of-2025 rosters (keeper pool).  Keeper_Analysis - every player's keeper cost (2025 price +$7; waivers $1+$7=$8), 2026 value, surplus.",False,11),
 ("Predicted_Keepers - each team's single best keeper (one-keeper rule).  League_Values - 2026 fair value (VOR on ESPN projections), league-expected price, and ESPN auction value side-by-side.",False,11),
 ("Draft_OverUnderpay - league pays ~fair for studs (0.98x), slight discount mid-tier (0.91x), big overpay on sleepers (2.61x). Pattern holds even excluding the most value-averse managers.",False,11),
 ("Value note: keeper cost = last year's price + $7 (waiver base $1). Values are VOR-based on ESPN 2026 projections; ESPN's raw auction values run low this early in preseason.",False,11),("",False,10),
 ("STRATEGY & DRAFT PLAN",True,12),
 ("Strategy_Success - what has correlated with winning (2023-25 auctions): draft RB-forward, work FAAB actively but save budget for late, don't get RB-trapped on waivers, avoid paying up for TE/QB.",False,11),
 ("Draft_Board - all 289 available players (post-keeper) with fair value, league-expected price, ESPN price, and a recommended max bid.",False,11),
 ("Scenarios - three ready-to-run auction plans for Ian's $142 (keep Puka): A) RB-Heavy (analytics pick), B) Two-Stud Hammer, C) Balanced Value Core.",False,11)]
for i,(t,b,sz) in enumerate(L,1):
    c=ab.cell(i,1); c.value=t; c.font=Font(bold=b,size=sz,color=navy if b else "000000"); c.alignment=Alignment(wrap_text=True,vertical="top")

money='#,##0.0'
add("Manager_AllTime","manager_alltime.csv",numfmts={7:'0.000',8:money,9:money,10:money,11:money,18:'0.0'})
add("Champions","champions.csv",widths={3:22,5:16})
add("Standings","standings.csv",numfmts={9:'0.000',13:money,14:money})
add("Luck_AllTime","luck_alltime.csv",numfmts={4:'0.0',5:'0.0'})
add("Luck_BySeason","luck_byseason.csv",numfmts={4:'0.00',5:'0.00',6:'0.000'})
add("H2H_Matrix","h2h_matrix.csv",widths={1:16})
add("Trades","trades.csv",widths={1:10,2:12,3:8,4:16,5:60})
add("Trades_ByManager","trades_bymanager.csv")
add("Draft_Fleaflicker","draft.csv",widths={6:26,7:24})
add("Draft_ESPN_Auction","espn_draft.csv",widths={4:18,5:18,6:14,7:24})
add("FAAB_Claims","faab_claims.csv",widths={3:14,5:24})
add("FAAB_ByManager","faab_by_manager.csv")
money2='#,##0.0'
add("Current_Rosters","current_rosters.csv")
add("Keeper_Analysis","keeper_analysis.csv",numfmts={5:'0.0',6:'0.0',7:'0.0'})
add("Predicted_Keepers","predicted_keepers.csv")
add("League_Values","league_adjusted_values.csv",widths={1:22},numfmts={3:'0.0',4:'0.0',5:'0.0',6:'0.0'})
add("Draft_OverUnderpay","overunderpay.csv",widths={1:22,5:34})
add("Strategy_Success","strategy_success.csv",widths={1:9,2:26,3:46,4:44})
add("Draft_Board","draft_board_2026.csv",widths={2:22})
add("Scenarios","scenarios.csv",widths={1:26,2:16,3:22,5:40})
wb.save("Gridiron_Grind_History.xlsx")
print("tabs:",wb.sheetnames)
